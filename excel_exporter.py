import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class ExcelExporter:
    def __init__(self, detector):
        self.detector = detector

    def parse_filename(self, filename):
        """
        Extracts name and ID from the filename.
        Assumes format: nameID_filename.py (e.g., johnDoe_12345.py)
        """
        try:
            name_id = filename.rsplit("_", 1)
            name = name_id[0]
            student_id = name_id[1].split(".")[0]
        except IndexError:
            name = filename
            student_id = "N/A"

        return name, student_id

    def export(self, file_path):
        try:
            # Create a new Excel workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Cheating Detection Report"

            # Setup headers
            headers = ["Name", "ID", "Actual Number", "Cheat (%)", "Final Grade"]
            sheet.append(headers)

            # Style headers
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Fill data
            report = self.detector.get_cheating_report()
            students_set = set()

            for line in report:
                # Extract the two filenames and the similarity score from the report
                try:
                    part1 = line.split(' between ')[1].split(' and ')
                    file1 = part1[0]
                    file2 = part1[1].split(' with ')[0]
                    similarity = float(line.split(' with a similarity score of ')[1]) * 100  # Convert to percentage
                except Exception as e:
                    print(f"Error parsing line: {line} - {e}")
                    continue

                # Extract name and ID
                name1, id1 = self.parse_filename(file1)
                name2, id2 = self.parse_filename(file2)

                # Append first student
                if (name1, id1) not in students_set:
                    final_grade1 = "0" if similarity == 100 else "Undetermined"
                    sheet.append([name1, id1, "", f"{similarity:.2f}", final_grade1])
                    students_set.add((name1, id1))

                # Append second student
                if (name2, id2) not in students_set:
                    final_grade2 = "0" if similarity == 100 else "Undetermined"
                    sheet.append([name2, id2, "", f"{similarity:.2f}", final_grade2])
                    students_set.add((name2, id2))

            # Add empty row for separation
            sheet.append([])

            # Add detailed report header
            detailed_report_header = ["Detailed Cheating Report"]
            sheet.append(detailed_report_header)
            sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=5)
            header_cell = sheet.cell(row=sheet.max_row, column=1)
            header_cell.font = Font(bold=True, color="FFFFFF")
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

            # Add detailed report information
            results = self.detector.get_detailed_results()
            for file1, file2, score in results:
                sheet.append([f'Possible cheating between {file1} and {file2} with a similarity score of {score:.2f}'])

            # Adjust column width
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except Exception as e:
                        print(f"Error adjusting column width: {e}")
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

            # Save the workbook
            workbook.save(file_path)
            print(f"Excel file saved successfully at {file_path}")

        except Exception as e:
            print(f"An error occurred while exporting to Excel: {e}")