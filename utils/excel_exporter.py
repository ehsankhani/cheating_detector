import openpyxl  # Importing the openpyxl library for working with Excel files
from openpyxl.styles import PatternFill, Font, Alignment  # Importing styles for formatting Excel cells
import pandas as pd  # Importing pandas for handling dataframes
import os


class ExcelExporter:
    def __init__(self, detector, folder_path):
        self.detector = detector  # Store the cheating detection object
        self.folder_path = folder_path

    def parse_filename(self, filename):
        """Extracts name and ID from the filename."""
        try:
            name_id = filename.rsplit("_", 1)
            name = name_id[0]
            student_id = name_id[1].split(".")[0]
        except IndexError:
            name = filename
            student_id = "N/A"
        return name, student_id

    def get_all_students(self):
        """Gets a list of all students' names and IDs from the folder path."""
        student_files = os.listdir(self.folder_path)
        all_students = {}

        for student_file in student_files:
            name, student_id = self.parse_filename(student_file)
            all_students[(name, student_id)] = 0  # Initialize cheat percentage as 0
        return all_students

    def export(self, output_file):
        # Placeholder dictionary for student scores
        students_scores = self.get_all_students()  # Initialize with all students

        # Collect detailed cheating report
        detailed_reports = []
        report = self.detector.get_cheating_report()

        for line in report:
            try:
                if 'Possible cheating between' in line and 'with an overall score of' in line:
                    parts = line.split(' between ')[1].split(' with an overall score of ')
                    files = parts[0].split(' and ')
                    file1, file2 = files[0].strip(), files[1].strip()

                    similarity = float(parts[1].split(' and ML prediction')[0].strip())
                    ml_prediction = 1

                    name1, id1 = self.parse_filename(file1)
                    name2, id2 = self.parse_filename(file2)

                    # Update cheating scores for the students involved in cheating
                    students_scores[(name1, id1)] = max(students_scores.get((name1, id1), 0), similarity * 100)
                    students_scores[(name2, id2)] = max(students_scores.get((name2, id2), 0), similarity * 100)

                    detailed_reports.append((file1, file2, similarity, ml_prediction))

            except Exception as e:
                print(f"Error parsing line: {line} - {e}")

        # Prepare summary data
        summary_data = []
        for (name, student_id), cheat_score in students_scores.items():
            final_grade = 0 if cheat_score == 100 else ''
            summary_data.append([name, student_id, '', cheat_score, final_grade])

        # Custom sorting function to handle numeric IDs
        def sort_key(item):
            # Extract the student ID
            student_id = item[1]
            try:
                # Try to convert the ID to an integer for sorting
                return int(student_id)
            except ValueError:
                # If it contains non-numeric characters, fallback to string sorting
                return student_id

        # Sort summary data based on the custom sorting key
        summary_data = sorted(summary_data, key=sort_key)

        # Convert sorted summary data to DataFrame
        summary_df = pd.DataFrame(summary_data, columns=['Name', 'ID', 'Actual Number', 'Cheat (%)', 'Final Grade'])

        # Prepare detailed report data
        detailed_pairs_data = []
        for file1, file2, score, ml_prediction in detailed_reports:
            detailed_pairs_data.append(
                f'Possible cheating between {file1} and {file2} '
                f'with an overall score of {score:.2f} and ML prediction: {ml_prediction}'
            )

        detailed_pairs_df = pd.DataFrame(detailed_pairs_data, columns=['Detailed Report'])

        # Write to Excel
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                summary_df.to_excel(writer, sheet_name='Summary', index=False)

                worksheet = writer.sheets['Summary']

                # Define fill styles for alternating colors
                light_blue_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
                dark_blue_fill = PatternFill(start_color='A4C8E1', end_color='A4C8E1', fill_type='solid')

                # Center align and format headers
                header_font = Font(name='Arial', size=12, bold=True)
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')

                # Format headers
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # Center align and format table content
                cell_font = Font(name='Arial', size=12)
                cell_alignment = Alignment(horizontal='center', vertical='center')

                # Apply styles to table rows
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                                                  min_col=1, max_col=summary_df.shape[1]), start=2):
                    for col_idx, cell in enumerate(row, start=1):
                        cell.font = cell_font
                        cell.alignment = cell_alignment
                        # Apply alternating fill colors
                        if row_idx % 2 == 0:
                            cell.fill = light_blue_fill
                        else:
                            cell.fill = dark_blue_fill

                # Write detailed pairs report below the summary
                startrow = len(summary_df) + 3
                detailed_pairs_df.to_excel(writer, sheet_name='Summary', startrow=startrow, index=False)

                for index in range(len(detailed_pairs_df)):
                    worksheet.merge_cells(start_row=startrow + index + 1, start_column=1,
                                          end_row=startrow + index + 1, end_column=summary_df.shape[1])
                    cell = worksheet.cell(row=startrow + index + 1, column=1)
                    cell.value = detailed_pairs_df.iloc[index, 0]
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                    # Apply alternating fill colors
                    if index % 2 == 0:
                        cell.fill = light_blue_fill
                    else:
                        cell.fill = dark_blue_fill

                    cell.font = Font(name='Arial', size=12, bold=False)

                # Adjust row height for better visibility
                for index in range(len(detailed_pairs_df)):
                    worksheet.row_dimensions[startrow + index + 1].height = 30

                # Add header style for the detailed report
                header_cell = worksheet.cell(row=startrow, column=1)
                header_cell.value = "Detailed Cheating Report"
                header_cell.font = Font(name='Arial', size=18, bold=True, color='FFFFFF')
                header_cell.fill = PatternFill(start_color='0530BB', end_color='0530BB', fill_type='solid')
                header_cell.alignment = Alignment(horizontal='center', vertical='center')

                worksheet.merge_cells(start_row=startrow, start_column=1, end_row=startrow,
                                      end_column=summary_df.shape[1])

                # Adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter  # Get the column letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception as e:
                            print(f"Error calculating max length for column {column_letter}: {e}")
                    adjusted_width = (max_length + 2)  # Adding extra space
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                last_index = startrow + len(detailed_pairs_df) + 1
                worksheet.merge_cells(start_row=last_index, start_column=1, end_row=last_index,
                                      end_column=summary_df.shape[1])
                last_cell = worksheet.cell(row=last_index, column=1)
                last_cell.value = "End of Detailed Cheating Report"
                last_cell.fill = PatternFill(start_color='0530BB', end_color='0530BB', fill_type='solid')
                last_cell.alignment = Alignment(horizontal='center', vertical='center')
                last_cell.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')

                print(f"Excel file saved successfully at {output_file}")

        except Exception as e:
            print(f"Error writing to Excel: {e}")

    def export_for_students(self, output_file):
        """Exports a simplified report with only name, ID, and final grade for all students, sorted by ID."""
        # Retrieve all students (assuming self.get_all_students() gives all students)
        students_scores = self.get_all_students()  # Initialize with all students having 0% cheating score

        # Collect detailed cheating report
        report = self.detector.get_cheating_report()

        for line in report:
            try:
                if 'Possible cheating between' in line and 'with an overall score of' in line:
                    parts = line.split(' between ')[1].split(' with an overall score of ')
                    files = parts[0].split(' and ')
                    file1, file2 = files[0].strip(), files[1].strip()

                    similarity = float(parts[1].split(' and ML prediction')[0].strip())

                    name1, id1 = self.parse_filename(file1)
                    name2, id2 = self.parse_filename(file2)

                    # Update cheating scores for students involved in cheating
                    students_scores[(name1, id1)] = max(students_scores.get((name1, id1), 0), similarity * 100)
                    students_scores[(name2, id2)] = max(students_scores.get((name2, id2), 0), similarity * 100)

            except Exception as e:
                print(f"Error parsing line: {line} - {e}")

        # Prepare student data for export
        student_data = []
        for (name, student_id), cheat_score in students_scores.items():
            final_grade = 0 if cheat_score == 100 else ''
            student_data.append([name, student_id, final_grade])

        # Custom sorting function to handle numeric IDs
        def sort_key(item):
            student_id = item[1]
            try:
                return int(student_id)  # Try to convert to int for sorting
            except ValueError:
                return student_id  # If not numeric, sort lexicographically

        # Sort students by ID
        student_data = sorted(student_data, key=sort_key)

        # Convert sorted student data to DataFrame
        student_df = pd.DataFrame(student_data, columns=['Name', 'ID', 'Final Grade'])

        # Write to Excel
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                student_df.to_excel(writer, sheet_name='Student Report', index=False)

                worksheet = writer.sheets['Student Report']

                # Define styles for headers
                header_font = Font(name='Arial', size=12, bold=True)
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')

                # Format headers
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # Center align and format table content
                cell_font = Font(name='Arial', size=12)
                cell_alignment = Alignment(horizontal='center', vertical='center')

                # Apply styles to table rows
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                                                  min_col=1, max_col=student_df.shape[1]), start=2):
                    for col_idx, cell in enumerate(row, start=1):
                        cell.font = cell_font
                        cell.alignment = cell_alignment

                # Adjust column widths for better readability
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter  # Get the column letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception as e:
                            print(f"Error calculating max length for column {column_letter}: {e}")
                    adjusted_width = (max_length + 2)  # Adding extra space
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                print(f"Student report saved successfully at {output_file}")

        except Exception as e:
            print(f"Error writing to Excel: {e}")
